# This is a python automation program using an excel file
#Melikaya Matiwane

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# loading our excel file
def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']

    # finds price values in the cells and take out 10% and store them on a new column
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # creates a barchart to displayed the fixed prices from column 4
    values = Reference(sheet,
      min_row=2,
      max_row=sheet.max_row,
      min_col=4,
      max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    workbook.save(filename)

    #The function can be reused to automate file changes

